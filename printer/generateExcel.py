import openpyxl
import os
import uuid

'''
    generateExcel
    功能： 通过传入的数据生成Excel 生成
        字段1 值1
        字段2 值2
    格式的Excel
    参数： data => 字段 - 数值 对应的json结构
    返回： Excel对应的路径
'''
def generateExcel(data) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    row = 1
    for key, value in data.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    file_path = os.path.join('./printer/tmp', str(uuid.uuid4()) + '.xlsx')
    
    wb.save(file_path)
    return os.path.abspath(file_path)


if __name__ == '__main__':
    #os.chdir('../')
    generateExcel({'cja': 1, 'jzh': 2})